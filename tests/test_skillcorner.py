"""SkillCorner ingest tests: sheet parsing and player matching, no real files."""

import openpyxl
import pandas as pd
import pytest
from sqlalchemy import create_engine

from lofc.ingest.skillcorner import (
    match_player_ids,
    player_season_records,
    team_season_records,
)

TEAM_HEADER = ["Team", "Team ID", "Season", "Season ID", "Minutes",
               "Count Performances (Physical Check passed)", "Distance P90", "PSV-99"]
PLAYER_HEADER = ["Player", "Player ID", "Birthdate", "Season", "Season ID", "Minutes",
                 "Count Performances (Physical Check passed)", "Distance P90", "PSV-99"]


@pytest.fixture
def workbook(tmp_path):
    """A minimal export with the two season sheets the loader reads."""
    import datetime
    book = openpyxl.Workbook()
    book.remove(book.active)
    team = book.create_sheet("Team x Season (avg P90)")
    team.append(TEAM_HEADER)
    team.append(["Leyton Orient FC", 2821, "2025/2026", 129, 92.8, 40, 9715.4, 28.4])
    team.append(["Wigan Athletic", 38, "2025/2026", 129, 92.1, 40, "null", 28.6])
    player = book.create_sheet("Player x Season (avg P90)")
    player.append(PLAYER_HEADER)
    player.append(["Dom Ballard", 666894, datetime.datetime(2005, 4, 1), "2025/2026", 129,
                   96.3, 39, 10201.5, 29.5])
    player.append(["Mystery Man", 999999, datetime.datetime(1990, 1, 1), "2025/2026", 129,
                   80.0, 5, 9000.0, 27.0])
    path = tmp_path / "SkillCorner-test.xlsx"
    book.save(path)
    return path


def test_team_records_parse_and_null_handling(workbook):
    records = team_season_records(workbook)
    assert len(records) == 2
    orient = records[0]
    assert orient["team_name"] == "Leyton Orient FC"
    assert orient["matches_measured"] == 40
    assert orient["distance_p90"] == 9715.4
    # SkillCorner writes the literal string 'null' for missing numbers.
    assert records[1]["distance_p90"] is None


def test_player_records_carry_iso_birth_dates(workbook):
    records = player_season_records(workbook)
    assert records[0]["player_name"] == "Dom Ballard"
    assert records[0]["birth_date"] == "2005-04-01"
    assert records[0]["psv99_kmh"] == 29.5


def test_match_player_ids_by_dob_and_name(workbook):
    records = player_season_records(workbook)
    engine = create_engine("sqlite://")
    pd.DataFrame([
        # Same DOB, slightly different name form: must match.
        {"player_id": 303856, "player_name": "Dominic Ballard", "birth_date": "2005-04-01"},
        # Same DOB as Mystery Man but a completely different name: must not match.
        {"player_id": 111, "player_name": "Someone Else Entirely", "birth_date": "1990-01-01"},
    ]).to_sql("players", engine, index=False)

    matched = match_player_ids(records, engine)

    assert matched == 1
    assert records[0]["player_id"] == 303856
    assert records[1]["player_id"] is None
